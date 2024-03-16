"""
Author：Lyon.Fan
File：F1.py
Date：2024/03/09
Note:
"""
from imp import reload
import numpy as np
import pandas as pd
import itertools
import openpyxl
import os
from typing import Tuple
from pprint import pprint

import Utility_F as u
reload(u)

def check_setting() -> None:
    """
        检查配置文件是否符合要求
    """
    list_stand_sheet: list[str] = ["其他字典",  "维度字典", "特征模板", "其他配置", "特征列表", "说明"]
    list_error_files: list[str] = []
    list_file = u.scaner_file(g_in_path)
    # 检查excel的sheet页面名称
    # ********************************
    for i_file in list_file:
        if "~" in i_file or not i_file.endswith(".xlsx"):
            continue
        for i_sheet in list_stand_sheet:
            try:
                pd.read_excel(f"{i_file}", sheet_name=f"{i_sheet}")
            except:
                list_error_files.append(i_file)
                break
        print(f"【{i_file}】sheet页面检查完毕")
    if len(list_error_files) > 0:
        print(f"XXXXXXXXXX【{i_file}】文件缺少【{i_sheet}】sheet页面")
        return False
    # 替换中文【：】
    # ********************************
    print("# ********************************")
    i_repalce = 0
    for i_file in list_file:
        if "~" in i_file or not i_file.endswith(".xlsx"):
            continue
        for i_sheet in list_stand_sheet:
            wb = openpyxl.load_workbook(f"{i_file}")  # 读取excel表格文件
            ws = wb[i_sheet]  # 选择key0工作表
            rows = ws.max_row  # 获取最大行
            columns = ws.max_column  # 获取最大列
            # 遍历表格内容,注意：openpyxl下标是从1开始
            for i in range(1, rows + 1):
                for j in range(1, columns + 1):
                    if str(ws.cell(i, j).value).find("：") < 0:
                        continue
                    ss = str(ws.cell(i, j).value).replace("：", ":")
                    ws.cell(i, j).value = ss
                    ws.cell(i, j).number_format = '@'
                    i_repalce += 1
            wb.save(f"{i_file}")
    print(f"【：】替换为【:】完毕, 共计替换了【{i_repalce}】个")    


def get_setting(p_file_name: str)-> Tuple[dict, dict, pd.DataFrame, pd.DataFrame, dict]:
    """
    根据设置自动生成变量中英文名称和SQL
    一个配置文件一个脚本、一个Excel结果文件
    :param p_file_name:配置文件文件名
    :return:
    """
    # ******************************************************************
    # 1、读取配置文件
    # ******************************************************************
    # 1.1、相关设置
    var_str_in_xlsx = f"{g_in_path}/{p_file_name}"  # 配置的xlsx全文件名
  
    # 1.2、读取模板、字典等配置表
    var_df_measurement = pd.read_excel(var_str_in_xlsx, sheet_name="其他字典")
    var_df_Dimension = pd.read_excel(var_str_in_xlsx, sheet_name="维度字典")
    temp_g_df_modle = pd.read_excel(var_str_in_xlsx, sheet_name="特征模板")
    temp_g_dict_others = pd.read_excel(var_str_in_xlsx, sheet_name="其他配置").set_index("项目名称")["项目值"].to_dict()
    temp_g_df_feature = pd.read_excel(var_str_in_xlsx, sheet_name="特征列表")
    # ******************************************************************
    # 2、设置总体字典，字典样式
    # {
    # '度量1': {'cname': '度量1',
    #         'ename': '度量1',
    #         'type': 'other',
    #         'values': {'交易金额': 'Calc1'}},
    # '时间间隔1': {'cname': '时间间隔1',
    #         'ename': '时间间隔1',
    #         'type': 'other',
    #         'values': {'1': '1',
    #                     '2': '2',
    #                     '3': '3',
    #                     '4': '4',
    #                     '5': '5',
    #                     '6': '6'}},
    # '转入传出标识': {'cname': '转入传出标识',
    #             'ename': 'flag_out_or_in',
    #             'type': 'Dim',
    #             'values': {'转入': '+', '转出': '-'}}
    # }
    # ******************************************************************
    temp_g_glob_dict = {}
    # 加入度量等字典
    for col in var_df_measurement:
        temp_values = {}
        for index, row in var_df_measurement.iterrows():
            if pd.notna(row[col]):
                key, value = row[col].split(':')
                temp_values[key] = value
        temp_g_glob_dict[col] = {'cname': col, 'ename': col, 'type':'other', 'values': temp_values}
    # 加入维度字典
    for col in var_df_Dimension:
        temp_canme,temp_ename = col.split(":")
        temp_values = {}
        for index, row in var_df_Dimension.iterrows():
            if pd.notna(row[col]):
                key, value = row[col].split(':')
                temp_values[key] = value
        temp_g_glob_dict[temp_canme] = {'cname': temp_canme, 'ename': temp_ename, 'type':'Dim', 'values': temp_values}
    # ******************************************************************
    return temp_g_glob_dict, temp_g_dict_others, temp_g_df_modle, temp_g_df_feature


def get_feature_df(var_glob_dict:dict, var_dict_others:dict, var_df_modle:pd.DataFrame, var_df_feature:pd.DataFrame):
    """
    根据模板和枚举取值，拼接成笛卡尔积，制作特征中文说明和特征英文名称
    Args:
        var_df_modle (pd.DataFrame): _description_
    Returns:
        _type_: _description_
    """
    for indexs in var_df_modle.index:
        str_model = var_df_modle.loc[indexs]["模板内容"]
        str_model_no = var_df_modle.loc[indexs]["模板编号"]
        str_category1 = var_df_modle.loc[indexs]["类别1"]
        str_category2 = var_df_modle.loc[indexs]["类别2"]
        str_category3 = var_df_modle.loc[indexs]["类别3"]
        str_calc = var_df_modle.loc[indexs]["计算类别"]
        no_param = 0  # 当前模板生成的特征的编号 从1开始 用于特征的英文名

        # 遍历得到要遍历的字段
        # 结果样式：['【时间字段】', '【大小额标识】', '【转入传出标识】', '【计算方法】', '【金额字段】']
        ls_param_name_1 = u.u_get_cell(str_model)
        # 去掉前后缀
        # 结果样式：['时间字段', '时间间隔', '大小额标识', '转入传出标识', '计算方法', '金额字段']
        ls_param_name_2 = []
        for i in range(len(ls_param_name_1)):
            ls_param_name_2.append(ls_param_name_1[i].strip(g_str_pre).strip(g_str_suf))   
        # 根据参数字段名，获得各个字段的枚举值，组成双层list，为后续笛卡尔积做准备
        # 结果样式：
        # [['时间字段'],[10, 15, 30, 60],['个人', '企业', '未知'],['转入', '转出'],['金额字段']]                 
        list_c_enum_for_descartes = []
        for i in ls_param_name_2:
            list_tmp = list(var_glob_dict[i]["values"].keys())
            list_c_enum_for_descartes.append(list_tmp)
        # 得到所有字典的笛卡尔积
        c = itertools.product(*list_c_enum_for_descartes)       
        # 将
        # 【时间字段】近【N】天，个人企业标识为【个人企业标识】，转入传出标识【转入传出标识】，的【计算方法】【度量】'
        # 转化为
        # 【{}】近【{}】天，个人企业标识为【{}】，转入传出标识【{}】，的【{}】【{}】
        str_feautre = str_model
        for i in ls_param_name_1:
            str_feautre = str_feautre.replace(i, g_str_pre + "{}" + g_str_suf)
        # 将
        # 【{}】近【{}】天，个人企业标识为【{}】，转入传出标识【{}】，的【{}】【{}】
        # 转化为
        # 【时间字段】近【720】天，个人企业标识为【个人】，转入传出标识【转出】，的【中位数】【金额字段】
        # 写入结果df
        for elem in c:
            no_param = no_param + 1
            # 英文名 样式:M005_Feat_0153
            str_tmp_1 = "{}_{}_{}".format(str_model_no, g_str_feature_flag, str(no_param).zfill(4))
            # 中文名
            str_tmp_2 = str_feautre.format(*elem)
            # 条件
            dict_condition = dict(zip(ls_param_name_2, elem))

            # 创建一个新的DataFrame来存储你想添加的数据
            new_row = pd.DataFrame([{
                '英文特征名称': str_tmp_1,
                "中文特征名称": str_tmp_2,
                "条件": dict_condition,
                "类别1": str_category1,
                "类别2": str_category2,
                "类别3": str_category3,
                "计算类别": str_calc
            }])
            # 使用pandas.concat来添加新行
            var_df_feature = pd.concat([var_df_feature, new_row], ignore_index=True)  
    return var_df_feature 


def get_sql(var_glob_dict:dict, var_dict_others:dict, var_df_modle:pd.DataFrame, var_df_feature:pd.DataFrame,p_file_name:str): 
    """
    根据模板生成指定的SQL 一个模板生成一个SQL文件 文件名是【模板编号.txt】
    """    
    list_single_feature_sql = [] # 单个特征的SQL
    list_sql_txt_name = [] # SQL语句所在txt的文件名
    list_tmp_table_list = [] # 所在的临时表表名
    
    for indexs, row in var_df_modle.iterrows():
        str_model_no = row["模板编号"]
        # 从特征中过滤当前这个模板的特征
        df_tmp_feature_list = var_df_feature[var_df_feature['英文特征名称'].str.startswith(str_model_no, na=False)] \
            .reset_index(drop=True)
        # 因为都是从一个模板中衍生出来的特征 所以取第一条 用来后续获取公共的条件 如时间字段、统计口径等
        dict_condition1 = dict(df_tmp_feature_list.loc[0]["条件"])

        str_sql_all_feature = ""
        # print(row["计算类别"])
        for indexs2, row2 in df_tmp_feature_list.iterrows(): 
            # 得到每个特征对应计算语句
            dict_condition = dict(row2["条件"])
            str_feature_ename = row2["英文特征名称"]
            # 维度字典作为最内层SQL
            str_sql_inner_calc = ""
            for key, value in dict_condition.items():
                if var_glob_dict[key]["type"] == "Dim":
                    str_sql_inner_calc = str_sql_inner_calc + u.u_get_space(8) + f" and ({var_glob_dict[key]['ename']}='{var_glob_dict[key]['values'][value]}') \n"
            # 添加时间间隔
            tmp_interval1 = int(dict_condition["时间间隔1"])
            tmp_calc1_part = dict_condition["度量1"]
            tmp_time_part1,tmp_time_part2 = u.get_date_interval(dict_condition["计算方法"],tmp_interval1)
            if tmp_time_part1 is not None:
                str_sql_date_calc1 = u.u_get_space(4) + "case( \n" \
                                    + u.u_get_space(8) + tmp_time_part1 + " \n" \
                                    + str_sql_inner_calc \
                                    + u.u_get_space(4) + f") when TRUE then {var_glob_dict['度量1']['values'][tmp_calc1_part]} else null end \n"
            if tmp_time_part2 is not None:
                str_sql_date_calc2 = u.u_get_space(4) + "case( \n" \
                                    + u.u_get_space(8) + tmp_time_part2 + " \n" \
                                    + str_sql_inner_calc \
                                    + u.u_get_space(4) + f") when TRUE then {var_glob_dict['度量1']['values'][tmp_calc1_part]} else null end \n"
            # 添加计算方法
            if tmp_time_part2 is None:
                str_sql_out_calc = u.u_get_clac_string(dict_condition["计算方法"]).format(str_sql_date_calc1, str_feature_ename)
            else:
                str_sql_out_calc = u.u_get_clac_string(dict_condition["计算方法"]).format(str_sql_date_calc1, str_sql_date_calc2, str_feature_ename)
            # 将计算方法每个特征的sql存在list中，便于后面添加到dataframe中
            list_single_feature_sql.append(str_sql_out_calc)
            # 将不同指标的SQL累加在一起
            str_sql_all_feature = "{}\n{}".format(str_sql_all_feature, str_sql_out_calc)
        
        # 基础日期字段 如消费日期
        dict(df_tmp_feature_list.loc[0]["条件"])
        null,col_time = u.u_get_dict("时间字段", dict_condition1["时间字段"], g_glob_dict)
        null,detla_time_unit = u.u_get_dict("时间间隔单位", dict_condition1["时间间隔单位"], g_glob_dict)
        # str_sql_date_delta
        str_sql_date_delta = u.u_get_date_delta_sql(delta_type=detla_time_unit
                                                , base_date_col=var_dict_others["对照日期"]
                                                , date_col=col_time)
        null, caliber_ename = u.u_get_dict("统计口径", dict_condition1["统计口径"], g_glob_dict)
        str_sql_model = f"drop table if exists {var_dict_others['中间输出表名']+ '_' + str_model_no};  \n" \
                        + f"create table {var_dict_others['中间输出表名']+ '_' + str_model_no} as \n" \
                        + f"select \n" \
                        + f"{caliber_ename}, {var_dict_others['对照日期']}" \
                        + str_sql_all_feature + "\n" \
                        + "from ( \n" \
                        + "     select *, \n" \
                        + f"    {str_sql_date_delta} \n" \
                        + f"     from {var_dict_others['基础表名']} \n" \
                        + f"     where {var_dict_others['过滤条件']} \n" \
                        + f") t \n" \
                        + f"group by {caliber_ename} ; \n"        
        # ******************************************************************
        # 根据模板生成指定的SQL 一个模板生成一个SQL
        # 目前设计的SQL基于中文字典
        # ******************************************************************
        str_sql_txt_name = f"{g_out_path}/out_{p_file_name.split('.')[0]}&{str_model_no}.txt" # 每个模板的每个公式一个txt文件
        with open(str_sql_txt_name, "w") as f:
            f.write(str_sql_model)  # 这句话自带文件关闭功能，不需要再写f.close()
        # SQL文件名和临时表名
        list_sql_txt_name.append(str_sql_txt_name)
        list_tmp_table_list.append(var_dict_others["中间输出表名"] + "_" + str_model_no)  
    var_df_feature["SQL_code"] = list_single_feature_sql

    var_df_modle["SQL_txt"] = list_sql_txt_name
    var_df_modle["tmp_table"] = list_tmp_table_list
    # ******************************************************************
    # 99、保存结果
    # ******************************************************************    
    write = pd.ExcelWriter(f"{g_out_path}/out_{p_file_name}" )
    var_df_modle.to_excel(excel_writer=write, sheet_name='特征模板', header=True,  index=False)
    var_df_feature.to_excel(excel_writer=write, sheet_name='特征列表', header=True,  index=False)
    write.close()    
    
    return var_df_modle, var_df_feature

def merge_txt_result(p_file_name:list[str]):
    merged_content = ""
    separator = "/* ************************************************* */ \n"
    # 合并所有字表txt的字表语句
    for i_in_excel_file in p_file_name: # 遍历配置文件名称
        list_out_sub_sql_file = u.scaner_file(g_out_path) # 遍历out文件夹下的所有文件
        for i_file in list_out_sub_sql_file:
            if i_file.startswith(f"{g_out_path}\\out_{i_in_excel_file.split('.')[0]}&") and i_file.endswith(".txt"): # 找到对应的txt文件
                with open(i_file, 'r') as file:
                    content = file.read()
                    # 将文件内容添加到merged_content字符串，每个文件内容后面都添加分隔符
                    merged_content += content + separator
    
    # 合并所有的字表合并为最终的表
    temp_all_table = g_df_all_modle["tmp_table"].tolist()
    str_1_talbe = temp_all_table[0]
    str_from_sql = f"from {str_1_talbe} \n"
    for i in temp_all_table[1:]:
        str_from_sql = str_from_sql + f"left join {i} on {str_1_talbe}.{g_dict_others['统计口径']} = {i}.{g_dict_others['统计口径']} \n"
        
    temp_all_col = g_df_all_feature["英文特征名称"].tolist()
    str_select_sql = f"select {str_1_talbe}.{g_dict_others['统计口径']} \n"
    for i in temp_all_col:
        str_select_sql = str_select_sql + f",{i} \n"
    str_all_sql = f"drop table if exists {g_dict_others['输出结果表名']}; \n" \
                + f"create table {g_dict_others['输出结果表名']} as \n" \
                + f"{str_select_sql}" \
                + f"{str_from_sql}" \
                + f"; \n"
    
    # 将合并的内容写入新的txt文件
    with open(f"{g_out_path}/out_merge_{p_file_name[0].split('.')[0]}.txt", 'w') as output_file:
        output_file.write(merged_content + str_all_sql)
    

def merge_excel_result(p_file_name:list[str]):
    write = pd.ExcelWriter(f"{g_out_path}/out_merge_{p_file_name[0].split('.')[0]}.xlsx" )
    g_df_all_modle.to_excel(excel_writer=write, sheet_name='特征模板', header=True,  index=False)
    g_df_all_feature.to_excel(excel_writer=write, sheet_name='特征列表', header=True,  index=False)
    write.close()   
    
    
# ***********************************************************************************
# ***********************************************************************************
# ***********************************************************************************
if __name__ == '__main__':
    """_summary_
        每个业务的配置文档 要放在单独的文件夹下
    """
    g_str_pre = "【"  # 前缀
    g_str_suf = "】"  # 后缀
    g_str_feature_flag = "Feat"  # 特征名称的前缀,特征名称样式为【M001Feat001】其中【M001】是模板的名称

    g_in_path = 'in'
    g_out_path = 'out'
    g_df_all_modle, g_df_all_feature = pd.DataFrame(), pd.DataFrame()
    g_list_file = ["表1_1P.xlsx","表2_1P.xlsx"]
    # 度量字典（度量等字典）、维度字典（维度字典）、其他配置字典（其他配置）
    g_glob_dict, g_dict_others = {}, {}    
    
    for i_file in g_list_file:
        # 特征模板（特征模板）,结果特征列表（特征列表）
        g_df_modle, g_df_feature = pd.DataFrame(), pd.DataFrame()
        print(f"****************************【{i_file}】：开始****************************")
        print(f"****************************【{i_file}】：配置文件检查****************************")
        check_setting()
        print(f"****************************【{i_file}】：读取配置文件 生成配置字典和配置模板****************************")
        g_glob_dict, g_dict_others, g_df_modle, g_df_feature = get_setting(i_file)
        print(f"****************************【{i_file}】：得到特征列表****************************")
        g_df_feature = get_feature_df(g_glob_dict, g_dict_others, g_df_modle, g_df_feature)
        print(f"****************************【{i_file}】：生成各个配置文件的特征SQL和结果文件****************************")
        g_df_modle, g_df_feature = get_sql(g_glob_dict, g_dict_others, g_df_modle, g_df_feature,i_file)
        print(f"****************************【{i_file}】：结果dataframe合并****************************")
        g_df_all_modle = pd.concat([g_df_all_modle, g_df_modle])
        g_df_all_feature = pd.concat([g_df_all_feature, g_df_feature])
    
    print(f"****************************合并所有txt和excel文件，并统计****************************")
    merge_txt_result(g_list_file)
    merge_excel_result(g_list_file)
    print(f"****************************完成！！！！！****************************")

